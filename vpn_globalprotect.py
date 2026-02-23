#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║  VPN SSL GlobalProtect — Automatización de Cuentas (INT/EXT)       ║
║  Genera credenciales + CLI PAN-OS + plantilla Excel                ║
║  Autor: Asistente IA Cybersecurity · Fecha: 2026-02-21             ║
╚══════════════════════════════════════════════════════════════════════╝

Modos de uso:
  1) Ticket desordenado:  python vpn_globalprotect.py ticket "texto..."
  2) Parámetros mínimos:  python vpn_globalprotect.py params --ritm ... --minsal ... --rut ... --tipo INT
  3) Desde Excel:         python vpn_globalprotect.py excel archivo.xlsx --row 2
  4) Generar plantilla:   python vpn_globalprotect.py template PlantillaVPN.xlsx
  5) Auto-tests:          python vpn_globalprotect.py --test
"""

import argparse
import csv
import json
import os
import re
import secrets
import string
import sys
from datetime import datetime, date

# ═══════════════════════════════════════════════════════════════════════
# FIX ENCODING EN WINDOWS (cp1252 no soporta emoji/unicode extendido)
# ═══════════════════════════════════════════════════════════════════════
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ═══════════════════════════════════════════════════════════════════════
# CONSTANTES / DEFAULTS
# ═══════════════════════════════════════════════════════════════════════

DEFAULTS = {
    "vsys": "vsys1",
    "zona_origen": "GP_VPN",
    "zona_destino": "inside_VPN",
    "grupo_int": "TIC_MINSAL_INT",
    "grupo_ext": "TIC_MINSAL_EXT",
    "log_setting": "Threat Level",
    "profile_group": "GRP_VPN",
    "crear_objetos": True,
    "incluir_commit": False,
    "schedule_start": None,          # None = hoy
}

USERNAME_LENGTH = 29
PASSWORD_LENGTH = 20

# Pool de caracteres para la password (sin espacios ni comillas)
_PW_UPPER   = string.ascii_uppercase
_PW_LOWER   = string.ascii_lowercase
_PW_DIGITS  = string.digits
_PW_SPECIAL = "!@#$%^&*()-_=+[]{}|;:,.<>?/"
_PW_ALL     = _PW_UPPER + _PW_LOWER + _PW_DIGITS + _PW_SPECIAL


# ═══════════════════════════════════════════════════════════════════════
# NORMALIZACIÓN
# ═══════════════════════════════════════════════════════════════════════

def normalize_minsal(raw: str) -> str:
    """
    Normaliza código MINSAL a formato 10 dígitos + '-' + sufijo.
    Ejemplo: '0806819-100' → '0000806819-100'
    """
    raw = raw.strip()
    if "-" not in raw:
        raise ValueError(f"MINSAL '{raw}' no contiene guion separador. Formato esperado: XXXXXXX-YYY")
    parts = raw.split("-", 1)
    num_part = re.sub(r"[^0-9]", "", parts[0])
    suffix = parts[1].strip()
    if not num_part:
        raise ValueError(f"MINSAL '{raw}': parte numérica vacía.")
    if not suffix:
        raise ValueError(f"MINSAL '{raw}': sufijo vacío.")
    num_padded = num_part.zfill(10)
    return f"{num_padded}-{suffix}"


def normalize_rut(raw: str) -> str:
    """
    Normaliza RUT chileno de forma hiper-robusta (ignora espacios, puntos, basura).
    Ejemplo: '1 6 0 4 7 8 7 3 - K' -> '16047873-k'
    """
    if not raw: return ""
    # 1. Limpiar todo lo que no sea dígito o k/K
    clean = re.sub(r"[^0-9kK-]", "", str(raw)).lower()
    
    if "-" not in clean:
        # Si no hay guion, intentamos ponerlo en la última posición (DV)
        if len(clean) >= 2:
            clean = clean[:-1] + "-" + clean[-1:]
        else:
            return clean # Muy corto para ser RUT
            
    parts = clean.rsplit("-", 1)
    num_part = re.sub(r"[^0-9]", "", parts[0])
    dv = parts[1].strip()[:1] # Solo el primer char después del guion
    
    if not num_part or not dv: return clean
    
    num_padded = num_part.zfill(8)
    return f"{num_padded}-{dv}"


def generate_username(minsal_norm: str, tipo: str, rut_norm: str) -> str:
    """
    Genera username VPN: <MINSAL_norm>_<INT|EXT>_<RUT_norm>
    Debe tener exactamente 29 caracteres.
    """
    tipo = tipo.upper()
    if tipo not in ("INT", "EXT"):
        raise ValueError(f"TIPO '{tipo}' inválido. Debe ser INT o EXT.")
    username = f"{minsal_norm}_{tipo}_{rut_norm}"
    if len(username) != USERNAME_LENGTH:
        raise ValueError(
            f"Username '{username}' tiene {len(username)} caracteres (esperado {USERNAME_LENGTH}).\n"
            f"  MINSAL_norm='{minsal_norm}' ({len(minsal_norm)} chars)\n"
            f"  TIPO='{tipo}' ({len(tipo)} chars)\n"
            f"  RUT_norm='{rut_norm}' ({len(rut_norm)} chars)\n"
            f"  Total = {len(minsal_norm)}+1+{len(tipo)}+1+{len(rut_norm)} = {len(username)}"
        )
    return username


def generate_password(length: int = 20, username: str = None) -> str:
    """
    Genera password criptográficamente segura de alta complejidad.
    Requisitos:
    - Longitud >= 16 (default 20)
    - >= 2 mayúsculas, >= 2 minúsculas, >= 2 dígitos, >= 1 especial.
    - Max 2 caracteres consecutivos repetidos.
    - No incluir el username (ni al derecho ni al revés).
    """
    if length < 16: length = 16
    
    # Pool de caracteres excluyendo potenciales problemas de escape
    upper = string.ascii_uppercase
    lower = string.ascii_lowercase
    digits = string.digits
    special = "!@#$^*()-_=+[]{}|;:,.<>?" # Removido % y & para evitar líos en algunos parsers
    all_chars = upper + lower + digits + special

    def is_valid(pw, user):
        if len(pw) < length: return False
        if sum(1 for c in pw if c in upper) < 2: return False
        if sum(1 for c in pw if c in lower) < 2: return False
        if sum(1 for c in pw if c in digits) < 2: return False
        if sum(1 for c in pw if c in special) < 1: return False
        
        # Max 2 caracteres consecutivos repetidos (aaa no permitido)
        for i in range(len(pw) - 2):
            if pw[i] == pw[i+1] == pw[i+2]: return False
            
        # No inclusion de username
        if user:
            u = str(user).lower()
            p = pw.lower()
            if u in p or u[::-1] in p: return False
        return True

    while True:
        # Asegurar mínimos
        pw_list = [
            secrets.choice(upper), secrets.choice(upper),
            secrets.choice(lower), secrets.choice(lower),
            secrets.choice(digits), secrets.choice(digits),
            secrets.choice(special)
        ]
        # Rellenar el resto
        pw_list += [secrets.choice(all_chars) for _ in range(length - len(pw_list))]
        
        # Mezclar
        for i in range(len(pw_list) - 1, 0, -1):
            j = secrets.randbelow(i + 1)
            pw_list[i], pw_list[j] = pw_list[j], pw_list[i]
        
        password = "".join(pw_list)
        if is_valid(password, username):
            return password


def extract_ips_cleaned(text: str) -> list:
    """
    Extrae IPv4 de forma hiper-robusta.
    """
    if not text: return []

    # 1. Heurística para IPs espaciadas: \d ( \d){0,2} limita octetos a 3 dígitos
    spaced_ip_pattern = re.compile(r"(\b\d(?:\s*\d){0,2}\s*\.\s*\d(?:\s*\d){0,2}\s*\.\s*\d(?:\s*\d){0,2}\s*\.\s*\d(?:\s*\d){0,2}\b)")
    
    # Pre-compactar el texto solo para encontrar IPs
    compacted_text = text
    for m in spaced_ip_pattern.finditer(text):
        original = m.group(1)
        compacted = original.replace(" ", "")
        compacted_text = compacted_text.replace(original, " " + compacted + " ")

    # 2. Regex IPv4 estricto: cada octeto 0-255
    _OCTET = r"(?:25[0-5]|2[0-4]\d|1?\d?\d)"
    IPV4_RE = re.compile(rf"(?<![.\d])({_OCTET}(?:\.{_OCTET}){{3}})(?![.\d])")

    # Limpiar envoltorios básicos
    t = compacted_text
    t = re.sub(r'https?://', " ", t, flags=re.IGNORECASE)
    t = re.sub(r'[\"\'()\[\]{}|?%&<>]', " ", t) # Más agresivo
    t = re.sub(r"[;,]+", " ", t)
    t = re.sub(r"\.{2,}", " ", t)

    raw_ips = IPV4_RE.findall(t)
    
    # Validación final
    valid_ips = []
    for ip in raw_ips:
        octets = ip.split(".")
        if all(0 <= int(o) <= 255 for o in octets):
            valid_ips.append(ip)

    unique_ips = list(dict.fromkeys(valid_ips))
    def ip_sort_key(ip_str): return tuple(int(o) for o in ip_str.split("."))
    return sorted(unique_ips, key=ip_sort_key)


def extract_ports_cleaned(text: str) -> list:
    """
    Extrae puertos de forma hiper-robusta.
    """
    if not text: return []
    
    # 1. De-spacing agresivo para bloques fragmentados
    t = re.sub(r"\b\d(?:\s+\d){1,4}\b", lambda m: m.group(0).replace(" ", ""), text)
    
    # 2. Prioridad a etiqueta
    results = []
    m_ports = re.search(r"(?i)(?:puertos?|dport|port|acceso|services?).*?[:=]+\s*([\d\s,;/\[\]-]+)", t)
    if m_ports:
        p_text = m_ports.group(1)
        for mp in re.finditer(r"(\d{2,5})", p_text):
            val = mp.group(1)
            if 80 <= int(val) <= 65535 and val not in ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027"]:
                results.append((val, "TCP"))
    
    # 3. Scan global solo si no se halló nada o para complementar (pero con máscara de IPs)
    if not results:
        ips = extract_ips_cleaned(t)
        for ip in ips: t = t.replace(ip, " [IP] ")
        
        # Limpiar ruidos
        t = re.sub(r'[><=!#$*~^|?%&(){}\[\]\\]', " ", t)
        
        for m in re.finditer(r"(\b\d{2,5}\b)", t):
            val = m.group(1)
            if 80 <= int(val) <= 65535 and val not in ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027", "777"]:
                results.append((val, "TCP"))

    # Eliminar duplicados y ordenar por número de puerto
    res = sorted(list(set(results)), key=lambda x: int(x[0]))
    return [f"{p}/{proto}" for p, proto in res]

# ═══════════════════════════════════════════════════════════════════════
# CLI RENDERING
# ═══════════════════════════════════════════════════════════════════════

def build_address_object_name(ip: str) -> str:
    """Convención de nombres para address objects: <IP>_32"""
    return f"{ip.strip()}_32"


def build_service_object_name(port_spec: str) -> str:
    """
    Convención de nombres para service objects: <PUERTO>_<PROTO>
    Acepta: '3389/TCP', '443/tcp', '3389_TCP', '53/UDP'
    """
    port_spec = port_spec.strip().upper()
    port_spec = port_spec.replace("/", "_")
    if "_" not in port_spec:
        # Asumir TCP si no se especifica protocolo
        port_spec = f"{port_spec}_TCP"
    return port_spec


def parse_port_spec(port_spec: str) -> tuple:
    """Parsea '3389/TCP' → ('3389', 'tcp')"""
    port_spec = port_spec.strip().upper().replace("_", "/")
    if "/" in port_spec:
        port, proto = port_spec.split("/", 1)
        return port.strip(), proto.strip().lower()
    return port_spec.strip(), "tcp"


def build_cli_int(params: dict) -> str:
    """Genera comandos CLI PAN-OS para tipo INT."""
    vsys = params.get("vsys", DEFAULTS["vsys"])
    vs = f"vsys {vsys}"  # Formato: "vsys vsys1"
    username = params["username"]
    password = params["password"]
    grupo = params.get("grupo_int", DEFAULTS["grupo_int"])

    lines = [
        "configure",
        f'set {vs} local-user-database user {username} password',
        "",
        password,
        password,
        "",
        f'set {vs} local-user-database user-group {grupo} user {username}',
    ]

    if params.get("incluir_commit", DEFAULTS["incluir_commit"]):
        lines += ["", "commit"]

    return "\n".join(lines)


def build_cli_ext(params: dict) -> str:
    """Genera comandos CLI PAN-OS para tipo EXT (con schedule, objetos, regla)."""
    vsys = params.get("vsys", DEFAULTS["vsys"])
    vs = f"vsys {vsys}"  # Formato: "vsys vsys1"
    username = params["username"]
    password = params["password"]
    grupo = params.get("grupo_ext", DEFAULTS["grupo_ext"])
    ritm = params["ritm"]
    vencimiento = params["vencimiento"]
    ips = params.get("ips", [])
    puertos = params.get("puertos", [])
    zona_origen = params.get("zona_origen", DEFAULTS["zona_origen"])
    zona_destino = params.get("zona_destino", DEFAULTS["zona_destino"])
    log_setting = params.get("log_setting", DEFAULTS["log_setting"])
    profile_group = params.get("profile_group", DEFAULTS["profile_group"])
    crear_objetos = params.get("crear_objetos", DEFAULTS["crear_objetos"])

    # Schedule
    start_date = params.get("schedule_start") or date.today().strftime("%Y/%m/%d")
    start_date_sched = start_date.replace("-", "/")
    end_date_sched = vencimiento.replace("-", "/")
    # Nombre del schedule en formato DD-MM-YYYY para coincidir con el ejemplo
    try:
        schedule_name = datetime.strptime(vencimiento, "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        schedule_name = vencimiento.replace("/", "-")

    lines = [
        "configure",
        f'set {vs} local-user-database user {username} password',
        "",
        password,
        password,
        "",
        f'set {vs} local-user-database user-group {grupo} user {username}',
        f'set {vs} schedule "{schedule_name}" schedule-type non-recurring {start_date_sched}@00:00-{end_date_sched}@23:59',
    ]

    # Address objects (en shared — crear solo si no existen)
    addr_obj_names = []
    if ips and crear_objetos:
        for ip in ips:
            ip = ip.strip()
            obj_name = build_address_object_name(ip)
            addr_obj_names.append(obj_name)
            lines.append(f'set shared address "{obj_name}" ip-netmask {ip}/32')
    elif ips:
        for ip in ips:
            addr_obj_names.append(build_address_object_name(ip.strip()))

    # Service objects (en shared — crear solo si no existen)
    svc_obj_names = []
    if puertos and crear_objetos:
        for ps in puertos:
            obj_name = build_service_object_name(ps)
            port, proto = parse_port_spec(ps)
            svc_obj_names.append(obj_name)
            lines.append(f'set shared service "{obj_name}" protocol {proto} port {port}')
    elif puertos:
        for ps in puertos:
            svc_obj_names.append(build_service_object_name(ps))

    # Destination y service strings para la regla (sin comillas entre items)
    dest_str = " ".join(addr_obj_names) if addr_obj_names else '"any"'
    svc_str = " ".join(svc_obj_names) if svc_obj_names else '"application-default"'

    # Regla de seguridad (Single Line)
    rule_name = f"{ritm}_policy"
    rule_cmd = (
        f'set {vs} rulebase security rules "{rule_name}" '
        f'description "{ritm}" from "{zona_origen}" to "{zona_destino}" '
        f'source "any" source-user [ "{username}" ] '
        f'destination [ {dest_str} ] application "any" service [ {svc_str} ] '
        f'action "allow" log-start no log-end yes log-setting "{log_setting}" '
        f'schedule "{schedule_name}" profile-setting group "{profile_group}"'
    )

    lines.append(rule_cmd)

    return "\n".join(lines)


def generate_ansible_playbook(params: dict) -> str:
    """
    Genera un Playbook de Ansible YAML usando la colección paloaltonetworks.panos.
    """
    ritm = params.get("ritm", "RITM_UNKNOWN")
    username = params.get("username", "user_unknown")
    vsys = params.get("vsys", "vsys1")
    
    lines = []
    lines.append("---")
    lines.append(f"- name: Configurar Acceso VPN - {ritm}")
    lines.append("  hosts: all")
    lines.append("  connection: local")
    lines.append("  gather_facts: false")
    lines.append("")
    lines.append("  vars:")
    lines.append("    # Credenciales del Firewall (Definir en inventario o vault)")
    lines.append("    provider:")
    lines.append("      ip_address: \"{{ device_ip }}\"")
    lines.append("      username: \"{{ device_user }}\"")
    lines.append("      password: \"{{ device_pass }}\"")
    lines.append("")
    lines.append("  tasks:")
    
    # Tarea 1: Asegurar el grupo de usuarios (INT o EXT)
    if params["tipo"] == "INT":
        grupo = params.get("grupo_int", "TIC_MINSAL_INT")
    else:
        grupo = params.get("grupo_ext", "TIC_MINSAL_EXT")
        
    lines.append(f"    - name: Asegurar Usuario en Grupo {grupo}")
    lines.append("      paloaltonetworks.panos.panos_type_cmd:")
    lines.append("        provider: \"{{ provider }}\"")
    lines.append(f"        xpath: \"/config/devices/entry[@name='localhost.localdomain']/vsys/entry[@name='{vsys}']/external-list/entry[@name='{grupo}']\"")
    lines.append(f"        element: \"<member>{username}</member>\"")
    lines.append("")

    # Si es EXT, crear objetos y reglas
    if params["tipo"] == "EXT":
        # Objetos de Dirección
        if params.get("crear_objetos") and params.get("ips"):
            for ip in params["ips"]:
                obj_name = build_address_object_name(ip)
                lines.append(f"    - name: Crear Address Object - {obj_name}")
                lines.append("      paloaltonetworks.panos.panos_address_object:")
                lines.append("        provider: \"{{ provider }}\"")
                lines.append(f"        name: \"{obj_name}\"")
                lines.append(f"        value: \"{ip}\"")
                lines.append(f"        description: \"GP VPN - {ritm}\"")
                lines.append(f"        vsys: \"{vsys}\"")
                lines.append("")

        # Objetos de Servicio
        if params.get("crear_objetos") and params.get("puertos"):
            for p_spec in params["puertos"]:
                svc_name = build_service_object_name(p_spec)
                p_val, p_proto = parse_port_spec(p_spec)
                lines.append(f"    - name: Crear Service Object - {svc_name}")
                lines.append("      paloaltonetworks.panos.panos_service_object:")
                lines.append("        provider: \"{{ provider }}\"")
                lines.append(f"        name: \"{svc_name}\"")
                lines.append(f"        protocol: \"{p_proto.lower()}\"")
                lines.append(f"        destination_port: \"{p_val}\"")
                lines.append(f"        vsys: \"{vsys}\"")
                lines.append("")

        # Regla de Seguridad
        lines.append(f"    - name: Crear Regla de Seguridad para {ritm}")
        lines.append("      paloaltonetworks.panos.panos_security_rule:")
        lines.append("        provider: \"{{ provider }}\"")
        lines.append(f"        rule_name: \"GP_{ritm}\"")
        lines.append(f"        source_zone: [\"{params.get('zona_origen')}\"]")
        lines.append(f"        destination_zone: [\"{params.get('zona_destino')}\"]")
        lines.append(f"        source_ip: [\"any\"]")
        lines.append(f"        destination_ip: {json.dumps([build_address_object_name(ip) for ip in params['ips']]) if params['ips'] else '[\"any\"]'}")
        lines.append(f"        service: {json.dumps([build_service_object_name(p) for p in params['puertos']]) if params['puertos'] else '[\"service-default\"]'}")
        lines.append(f"        application: [\"any\"]")
        lines.append(f"        action: \"allow\"")
        lines.append(f"        vsys: \"{vsys}\"")
        lines.append(f"        description: \"Acceso GP VPN - {ritm}\"")
        if params.get("log_setting"):
            lines.append(f"        log_setting: \"{params['log_setting']}\"")
        if params.get("profile_group"):
            lines.append(f"        group_profile: \"{params['profile_group']}\"")
        lines.append("")

    # Commit opcional
    if params.get("incluir_commit"):
        lines.append("    - name: Commit Changes")
        lines.append("      paloaltonetworks.panos.panos_commit_firewall:")
        lines.append("        provider: \"{{ provider }}\"")
        lines.append("")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════
# PARSEO DE TICKET ROBUSTO (MODO 1)
# ═══════════════════════════════════════════════════════════════════════

TICKET_SYNONYMS = {
    "ritm": ["ritm", "requerimiento", "req", "ticket", "solicitud", "n° req", "id"],
    "minsal": ["numero minsal", "n° minsal", "minsal", "cod minsal", "id minsal"],
    "rut": ["rut", "run"],
    "tipo": ["seguridad", "tipo", "canal", "acceso"],
    "vencimiento": ["fecha expiracion", "expira", "vigencia", "vence", "hasta", "expiry"],
    "nombre": ["nombre completo", "nombre", "solicitante", "usuario", "beneficiario", "titular"],
    "email": ["email", "correo", "e-mail", "mail"],
    "ips": ["listado de ip", "ips", "ip a acceder", "destinos", "redes", "hosts"],
    "puertos": ["puertos", "ports", "dport", "puerto destino", "puerto"],
    "obs": ["observación", "observacion", "obs", "notas", "comentario"],
    "hospital": ["nombre hospital", "hospital", "establecimiento", "centro", "unidad", "departamento"],
    "region": ["región", "region", "reg", "zona"],
    "comuna": ["comuna", "municipio", "ciudad", "localidad"],
    "contratante": ["contratante", "cliente", "servicio de salud", "ss", "mandante"]
}

SECTION_HEADERS = {
    "solicitante": ["datos solicitante", "solicitante", "requester", "peticionario"],
    "beneficiario": ["datos beneficiarios", "beneficiario", "usuario", "titular", "acceso para", "beneficiario/a"]
}

def normalize_ticket_text(text: str) -> str:
    """
    Limpieza agresiva para búsqueda de etiquetas.
    - Quita símbolos de adorno.
    - Colapsa espacios.
    - Compacta caracteres individuales separados por espacios (De-Spacing).
    """
    if not text: return ""
    
    # Limpieza básica: símbolos a espacios, colapsar espacios.
    t = re.sub(r'[><=!#$*~^|?%&(){}\[\]\\]', " ", text)
    t = re.sub(r"\s+", " ", t)
    return t.strip()
    
    # 3. Colapsar puntos múltiples y espacios horizontales
    t = re.sub(r"\.{2,}", " ", t)
    t = re.sub(r"[ \t]+", " ", t)
    
    return t.strip()

def parse_ticket(text: str) -> dict:
    """
    Extracción hiper-robusta final de última generación.
    Combina 'Space-Agnostic Regex' con 'Section-Aware Blocks'.
    """
    if not text: return {}
    
    data = {}
    
    # helper de-spacer
    def ds(s): return re.sub(r"\s+", "", s)
    txt_ds = ds(text)

    # 1. IDENTIFICADORES GLOBALES (Prioridad Máxima en texto sin espacios)
    ritms = re.findall(r"RITM(\d{5,})", txt_ds, re.I)
    if ritms:
        data["ritm"] = f"RITM{max(ritms, key=len)}".upper()
    else:
        m_req = re.search(r"(\d{4}-\d{4})", txt_ds)
        if m_req: data["ritm"] = m_req.group(1)
        
    m_min = re.search(r"(\d{5,10}-\d{3})", txt_ds)
    if m_min: data["minsal"] = m_min.group(1)

    # 2. BLOQUES Y DATOS DE BENEFICIARIO
    # Encontrar inicio de beneficiario
    idx_ben = -1
    for syn in SECTION_HEADERS["beneficiario"]:
        m = re.search(rf"(?i)(?:^|\s){re.escape(syn)}s?(?:[^\w]|$)", text)
        if m and (idx_ben == -1 or m.start() < idx_ben):
            idx_ben = m.start()
    
    ben_block = text[idx_ben:] if idx_ben != -1 else text
    ben_ds = ds(ben_block)

    # RUT (en bloque beneficiario)
    m_rut = re.search(r"RUT[:=]*(\d{1,2}\.?\d{3}\.?\d{3}-?[\dkK])", ben_ds, re.I)
    if not m_rut: m_rut = re.search(r"(\d{1,2}\.?\d{3}\.?\d{3}-?[\dkK])", ben_ds, re.I)
    if m_rut: data["rut"] = normalize_rut(m_rut.group(1))

    # Nombre (en bloque beneficiario)
    m_name = re.search(r"(?i)(?:nombre completo|nombre|beneficiario).*?[:=]+\s*([a-zA-Z\s]{3,40})(?=\s(?:rut|email|cargo|fono|establecimiento|minsal)|$|;|\n|---)", ben_block)
    if m_name:
        val = m_name.group(1).strip()
        data["nombre"] = re.sub(r"\s{2,}", " ", val).strip()

    # IPs y Puertos (Global)
    data["ips"] = []
    ip_pat = r"(\b\d(?:\s*\d){0,2}(?:\s*\.\s*\d(?:\s*\d){0,2}){3}\b)"
    for m in re.finditer(ip_pat, text):
        val = ds(m.group(1))
        if all(0 <= int(o) <= 255 for o in val.split(".")):
            data["ips"].append(val)
    data["ips"] = sorted(list(set(data["ips"])))
    
    # 2.5 Puertos (Prioridad a etiquetas para evitar ruido de IPs/RUTs)
    data["puertos"] = []
    # Buscar patrón de etiqueta
    m_ports = re.search(r"(?i)(?:puertos?|dport|port|acceso|services?).*?[:=]+\s*([\d\s,;/\[\]-]+)", text)
    if m_ports:
        p_text = m_ports.group(1)
        for mp in re.finditer(r"(\d(?:\s*\d){1,4})", p_text):
            val = ds(mp.group(1))
            if 80 <= int(val) <= 65535 and val not in ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027"]:
                data["puertos"].append(f"{val}/TCP")
    
    # Fallback global solo si no se detectó nada con etiquetas
    if not data["puertos"]:
        # Limpiar IPs del texto para evitar falsos positivos
        txt_masked = text
        for ip in data["ips"]: txt_masked = txt_masked.replace(ip, " [IP] ")
        
        for mp in re.finditer(r"\b(\d(?:\s*\d){1,4})\b", txt_masked):
            val = ds(mp.group(1))
            if 80 <= int(val) <= 65535 and val not in ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027", "777"]:
                data["puertos"].append(f"{val}/TCP")

    data["puertos"] = sorted(list(set(data["puertos"])), key=lambda x: int(x.split("/")[0]))

    # 3. Datos complementarios
    m_v = re.search(r"(?i)(?:vencimiento|expiracion|hasta).*?[:=>]+\s*(\d(?:\s*\d){3}-\d(?:\s*\d)-\d(?:\s*\d))", text)
    if m_v: data["vencimiento"] = ds(m_v.group(1))
    else:
        m_v2 = re.search(r"(\d{4}-\d{2}-\d{2})", txt_ds)
        if m_v2: data["vencimiento"] = m_v2.group(1)
        
    # Email (Priorizar beneficiario)
    m_em = re.search(r"([a-zA-Z0-9._%+-]+\s*@\s*[a-zA-Z0-9.-]+\s*\.\s*[a-zA-Z]{2,})", ben_block)
    if not m_em: m_em = re.search(r"([a-zA-Z0-9._%+-]+\s*@\s*[a-zA-Z0-9.-]+\s*\.\s*[a-zA-Z]{2,})", text)
    if m_em: data["email"] = ds(m_em.group(1)).lower()

    if data.get("vencimiento"): data["tipo"] = "EXT"
    else: data["tipo"] = "INT" if "INT" in text.upper() else "EXT"

    return data


# ═══════════════════════════════════════════════════════════════════════
# LECTURA DESDE EXCEL (MODO 3)
# ═══════════════════════════════════════════════════════════════════════

def read_excel_row(filepath: str, row: int = 2) -> dict:
    """Lee una fila de la plantilla Excel y retorna dict con los campos."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Se requiere 'openpyxl'. Instala con: pip install openpyxl")

    wb = load_workbook(filepath, data_only=True)
    ws = wb["SolicitudesVPN"]

    # Columnas esperadas: A=RITM, B=MINSAL, C=RUT, D=TIPO, E=VENCIMIENTO, F=IPS, G=PUERTOS, H=OBS
    data = {}
    val = lambda col: ws[f"{col}{row}"].value

    data["ritm"] = str(val("A") or "").strip()
    data["minsal"] = str(val("B") or "").strip()
    data["rut"] = str(val("C") or "").strip()
    data["tipo"] = str(val("D") or "").strip().upper()

    venc = val("E")
    if isinstance(venc, datetime):
        data["vencimiento"] = venc.strftime("%Y-%m-%d")
    elif isinstance(venc, date):
        data["vencimiento"] = venc.strftime("%Y-%m-%d")
    elif venc:
        data["vencimiento"] = str(venc).strip()

    ips_raw = str(val("F") or "").strip()
    if ips_raw:
        data["ips"] = [ip.strip() for ip in re.split(r"[,;\n]+", ips_raw) if ip.strip()]

    puertos_raw = str(val("G") or "").strip()
    if puertos_raw:
        data["puertos"] = [p.strip() for p in re.split(r"[,;\n]+", puertos_raw) if p.strip()]

    data["obs"] = str(val("H") or "").strip()

    wb.close()
    return data


# ═══════════════════════════════════════════════════════════════════════
# GENERACIÓN DE PLANTILLA EXCEL
# ═══════════════════════════════════════════════════════════════════════

def create_template_excel(filepath: str) -> None:
    """Genera la plantilla Excel con validaciones y fórmulas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SolicitudesVPN"

    # ── Estilos ──
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    calc_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    calc_font = Font(name="Calibri", bold=True, color="2F5496", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ── Headers ──
    headers = [
        ("A", "RITM",         18, "Obligatorio"),
        ("B", "MINSAL",       20, "Obligatorio"),
        ("C", "RUT",          16, "Obligatorio"),
        ("D", "TIPO",         8,  "INT o EXT"),
        ("E", "VENCIMIENTO",  16, "YYYY-MM-DD"),
        ("F", "IPS",          30, "Solo EXT, separar con coma"),
        ("G", "PUERTOS",      25, "Solo EXT, ej: 3389/TCP"),
        ("H", "OBS",          35, "Observación"),
        ("I", "MINSAL_norm",  20, "← Fórmula automática"),
        ("J", "RUT_norm",     16, "← Fórmula automática"),
        ("K", "USERNAME",     35, "← Fórmula automática"),
        ("L", "LargoUsername", 16, "← Debe ser 29"),
    ]

    for col_letter, name, width, comment in headers:
        cell = ws[f"{col_letter}1"]
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[col_letter].width = width

    # ── Comentarios en fila 2 (guía) ──
    for col_letter, name, width, comment in headers:
        cell = ws[f"{col_letter}2"]
        cell.value = comment
        cell.font = Font(italic=True, color="808080", size=9)
        cell.alignment = Alignment(horizontal="center")

    # ── Fórmulas filas 3-102 ──
    for row in range(3, 103):
        # I: MINSAL_norm = REPT("0",10-LEN(LEFT(B,FIND("-",B)-1))) & B
        ws[f"I{row}"] = (
            f'=IF(B{row}="","",REPT("0",10-LEN(LEFT(B{row},FIND("-",B{row})-1)))&B{row})'
        )
        ws[f"I{row}"].fill = calc_fill
        ws[f"I{row}"].font = calc_font

        # J: RUT_norm — quita puntos, padding 8 dígitos
        # Simplificación: asume input sin puntos, solo padding
        ws[f"J{row}"] = (
            f'=IF(C{row}="","",REPT("0",8-LEN(LEFT(SUBSTITUTE(C{row},".",""),FIND("-",SUBSTITUTE(C{row},".","")))))'
            f'&SUBSTITUTE(C{row},".",""))'
        )
        # Fórmula más simple y robusta para RUT:
        ws[f"J{row}"] = (
            f'=IF(C{row}="","",LET(clean,SUBSTITUTE(C{row},".",""),'
            f'numpart,LEFT(clean,FIND("-",clean)-1),'
            f'dvpart,MID(clean,FIND("-",clean),2),'
            f'REPT("0",8-LEN(numpart))&numpart&dvpart))'
        )
        ws[f"J{row}"].fill = calc_fill
        ws[f"J{row}"].font = calc_font

        # K: USERNAME = MINSAL_norm & "_" & TIPO & "_" & RUT_norm
        ws[f"K{row}"] = (
            f'=IF(OR(I{row}="",D{row}="",J{row}=""),"",I{row}&"_"&D{row}&"_"&LOWER(J{row}))'
        )
        ws[f"K{row}"].fill = calc_fill
        ws[f"K{row}"].font = calc_font

        # L: LargoUsername
        ws[f"L{row}"] = f'=IF(K{row}="","",LEN(K{row}))'
        ws[f"L{row}"].fill = calc_fill
        ws[f"L{row}"].font = calc_font

    # ── Validación TIPO: solo INT o EXT ──
    dv_tipo = DataValidation(
        type="list",
        formula1='"INT,EXT"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Tipo inválido",
        error="Solo se permite INT o EXT"
    )
    dv_tipo.sqref = "D3:D102"
    ws.add_data_validation(dv_tipo)

    # ── Formato condicional en L (LargoUsername) ──
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "L3:L102",
        CellIsRule(operator="equal", formula=["29"], fill=green_fill)
    )
    ws.conditional_formatting.add(
        "L3:L102",
        CellIsRule(operator="notEqual", formula=["29"], fill=red_fill)
    )

    # ── Freeze panes ──
    ws.freeze_panes = "A3"

    # ── Guardar ──
    wb.save(filepath)
    print(f"✅ Plantilla Excel creada: {filepath}")


# ═══════════════════════════════════════════════════════════════════════
# EXPORT CSV / EXCEL
# ═══════════════════════════════════════════════════════════════════════

def export_csv(params: dict, filepath: str) -> None:
    """Exporta los datos normalizados a CSV."""
    fieldnames = [
        "ritm", "minsal_input", "minsal_norm", "rut_input", "rut_norm",
        "tipo", "username", "vencimiento", "ips", "puertos", "obs"
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        row = {k: params.get(k, "") for k in fieldnames}
        if isinstance(row.get("ips"), list):
            row["ips"] = ",".join(row["ips"])
        if isinstance(row.get("puertos"), list):
            row["puertos"] = ",".join(row["puertos"])
        writer.writerow(row)
    print(f"✅ CSV exportado: {filepath}")


def export_excel_output(params: dict, filepath: str) -> None:
    """Exporta los datos normalizados a Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    headers = ["RITM", "MINSAL_input", "MINSAL_norm", "RUT_input", "RUT_norm",
               "TIPO", "USERNAME", "PASSWORD", "VENCIMIENTO", "IPs", "PUERTOS", "OBS"]
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    values = [
        params.get("ritm", ""),
        params.get("minsal_input", ""),
        params.get("minsal_norm", ""),
        params.get("rut_input", ""),
        params.get("rut_norm", ""),
        params.get("tipo", ""),
        params.get("username", ""),
        params.get("password", ""),
        params.get("vencimiento", ""),
        ",".join(params.get("ips", [])),
        ",".join(params.get("puertos", [])),
        params.get("obs", ""),
    ]
    for col, v in enumerate(values, 1):
        ws.cell(row=2, column=col, value=v)

    wb.save(filepath)
    print(f"✅ Excel exportado: {filepath}")


# ═══════════════════════════════════════════════════════════════════════
# RENDER PRINCIPAL DE SALIDA
# ═══════════════════════════════════════════════════════════════════════

def render_output(params: dict) -> str:
    """Genera la salida formateada completa."""
    lines = []
    sep = "═" * 70

    # ── 1. DATOS NORMALIZADOS ──
    lines.append(f"\n{sep}")
    lines.append("  📋 DATOS NORMALIZADOS")
    lines.append(sep)
    table_data = {
        "RITM":           params.get("ritm", "N/A"),
        "MINSAL (input)": params.get("minsal_input", "N/A"),
        "MINSAL (norm)":  params.get("minsal_norm", "N/A"),
        "RUT (input)":    params.get("rut_input", "N/A"),
        "RUT (norm)":     params.get("rut_norm", "N/A"),
        "TIPO":           params.get("tipo", "N/A"),
        "USERNAME":       f'{params.get("username", "N/A")}  ({len(params.get("username", ""))} chars)',
        "VENCIMIENTO":    params.get("vencimiento", "N/A"),
        "IPs destino":    ", ".join(params.get("ips", [])) or "N/A",
        "Puertos":        ", ".join(params.get("puertos", [])) or "N/A",
        "vsys":           params.get("vsys", DEFAULTS["vsys"]),
        "Zona origen":    params.get("zona_origen", DEFAULTS["zona_origen"]),
        "Zona destino":   params.get("zona_destino", DEFAULTS["zona_destino"]),
        "Grupo":          params.get("grupo_int" if params.get("tipo") == "INT" else "grupo_ext",
                                     DEFAULTS["grupo_int"] if params.get("tipo") == "INT" else DEFAULTS["grupo_ext"]),
    }
    max_key = max(len(k) for k in table_data)
    for k, v in table_data.items():
        lines.append(f"  {k:<{max_key + 2}} │ {v}")

    # ── JSON ──
    lines.append(f"\n  📦 JSON:")
    json_data = {k.replace(" ", "_").replace("(", "").replace(")", ""): v
                 for k, v in table_data.items()}
    lines.append(json.dumps(json_data, indent=2, ensure_ascii=False))

    # ── 2. CREDENCIALES ──
    lines.append(f"\n{sep}")
    lines.append("  🔐 CREDENCIALES")
    lines.append(sep)
    lines.append(f"  USERNAME:  {params['username']}")
    lines.append(f"  PASSWORD:  {params['password']}")
    lines.append(f"  (Password: {len(params['password'])} chars)")

    # ── 3. COMANDOS CLI ──
    lines.append(f"\n{sep}")
    lines.append("  💻 COMANDOS CLI PAN-OS (COPY/PASTE)")
    lines.append(sep)
    if params["tipo"] == "INT":
        cli = build_cli_int(params)
    else:
        cli = build_cli_ext(params)
    lines.append(cli)

    # ── 4. ANSIBLE PLAYBOOK ──
    if params.get("ansible_playbook"):
        lines.append(f"\n{sep}")
        lines.append("  🛠️  ANSIBLE PLAYBOOK (YAML)")
        lines.append(sep)
        lines.append(params["ansible_playbook"])

    lines.append(f"\n{sep}")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════
# PROCESAMIENTO CENTRAL
# ═══════════════════════════════════════════════════════════════════════

def process_data(data: dict) -> dict:
    """
    Toma datos crudos, normaliza, genera credenciales y retorna params completos.
    Usa placeholders si faltan datos para no bloquear la generación del CLI.
    """
    # Valores por defecto / Placeholders
    ritm = data.get("ritm") or "[FALTA_RITM]"
    minsal_raw = data.get("minsal") or "0000000-000"
    rut_raw = data.get("rut") or "00000000-0"
    tipo = data.get("tipo", "").upper() or "EXT"
    vencimiento = data.get("vencimiento") or "2099-12-31"

    # Intentar normalización
    try: minsal_norm = normalize_minsal(minsal_raw)
    except: minsal_norm = "0000000000-000"
    
    try: rut_norm = normalize_rut(rut_raw)
    except: rut_norm = "00000000-0"

    # Generar Username/Password (siempre debe funcionar con los defaults arriba)
    try:
        username = generate_username(minsal_norm, tipo, rut_norm)
    except:
        username = f"{minsal_norm[:10]}_{tipo}_{rut_norm}"

    password = generate_password(username=username)

    # IPs y Puertos (Solo extraer si no vienen ya listos)
    ips_raw = data.get("ips", [])
    if isinstance(ips_raw, list) and len(ips_raw) > 0:
        # Ya vienen extraídos y limpios de parse_ticket
        ips_cleaned = sorted(list(set(ips_raw)))
    else:
        if isinstance(ips_raw, list): ips_raw = " ".join(ips_raw)
        ips_cleaned = extract_ips_cleaned(str(ips_raw))

    puertos_raw = data.get("puertos", [])
    if isinstance(puertos_raw, list) and len(puertos_raw) > 0:
        # Ya vienen extraídos
        puertos_cleaned = sorted(list(set(puertos_raw)))
    else:
        if isinstance(puertos_raw, list): puertos_raw = " ".join(puertos_raw)
        puertos_cleaned = extract_ports_cleaned(str(puertos_raw))
    params = {
        "ritm": ritm.upper(),
        "minsal_input": minsal_raw,
        "minsal_norm": minsal_norm,
        "rut_input": rut_raw,
        "rut_norm": rut_norm,
        "tipo": tipo,
        "username": username,
        "password": password,
        "vencimiento": vencimiento,
        "ips": ips_cleaned,
        "puertos": puertos_cleaned,
        "obs": data.get("obs", ""),
        # Defaults
        "vsys": data.get("vsys", DEFAULTS["vsys"]),
        "zona_origen": data.get("zona_origen", DEFAULTS["zona_origen"]),
        "zona_destino": data.get("zona_destino", DEFAULTS["zona_destino"]),
        "grupo_int": data.get("grupo_int", DEFAULTS["grupo_int"]),
        "grupo_ext": data.get("grupo_ext", DEFAULTS["grupo_ext"]),
        "log_setting": data.get("log_setting", DEFAULTS["log_setting"]),
        "profile_group": data.get("profile_group", DEFAULTS["profile_group"]),
        "crear_objetos": data.get("crear_objetos", DEFAULTS["crear_objetos"]),
        "incluir_commit": data.get("incluir_commit", DEFAULTS["incluir_commit"]),
        "schedule_start": data.get("schedule_start", DEFAULTS["schedule_start"]),
    }

    # Generar Playbook Ansible
    params["ansible_playbook"] = generate_ansible_playbook(params)

    return params


# ═══════════════════════════════════════════════════════════════════════
# SELF-TESTS
# ═══════════════════════════════════════════════════════════════════════

def run_tests():
    """Suite de auto-tests integrados."""
    passed = 0
    failed = 0
    total = 0

    def check(name, condition, detail=""):
        nonlocal passed, failed, total
        total += 1
        if condition:
            print(f"  ✅ {name}")
            passed += 1
        else:
            print(f"  ❌ {name} — {detail}")
            failed += 1

    print("\n🧪 Ejecutando tests...\n")

    # Test 1: MINSAL normalización
    check("MINSAL norm básica",
          normalize_minsal("0806819-100") == "0000806819-100",
          f"obtenido: {normalize_minsal('0806819-100')}")
    check("MINSAL norm ya completo",
          normalize_minsal("0000806819-100") == "0000806819-100")
    check("MINSAL norm corto",
          normalize_minsal("12345-200") == "0000012345-200")
    check("MINSAL norm largo",
          len(normalize_minsal("0806819-100")) == 14,
          f"largo: {len(normalize_minsal('0806819-100'))}")

    # Test 2: RUT normalización
    check("RUT norm básica",
          normalize_rut("12994496-k") == "12994496-k",
          f"obtenido: {normalize_rut('12994496-k')}")
    check("RUT norm con puntos",
          normalize_rut("12.994.496-K") == "12994496-k",
          f"obtenido: {normalize_rut('12.994.496-K')}")
    check("RUT norm padding",
          normalize_rut("1234567-8") == "01234567-8",
          f"obtenido: {normalize_rut('1234567-8')}")
    check("RUT norm largo",
          len(normalize_rut("12994496-k")) == 10,
          f"largo: {len(normalize_rut('12994496-k'))}")

    # Test 3: Username
    un = generate_username("0000806819-100", "INT", "12994496-k")
    check("Username INT valor",
          un == "0000806819-100_INT_12994496-k",
          f"obtenido: {un}")
    check("Username INT largo 29",
          len(un) == 29,
          f"largo: {len(un)}")

    un_ext = generate_username("0000806819-100", "EXT", "12994496-k")
    check("Username EXT valor",
          un_ext == "0000806819-100_EXT_12994496-k",
          f"obtenido: {un_ext}")
    check("Username EXT largo 29",
          len(un_ext) == 29,
          f"largo: {len(un_ext)}")

    # Test 4: Password
    for i in range(5):
        pw = generate_password(length=20, username="testadmin")
        check(f"Password #{i+1} largo 20",
              len(pw) == 20,
              f"largo: {len(pw)}")
        check(f"Password #{i+1} >= 2 mayús",
              sum(1 for c in pw if c in string.ascii_uppercase) >= 2)
        check(f"Password #{i+1} >= 2 minús",
              sum(1 for c in pw if c in string.ascii_lowercase) >= 2)
        check(f"Password #{i+1} >= 2 cifras",
              sum(1 for c in pw if c in string.digits) >= 2)
        check(f"Password #{i+1} >= 1 esp",
              sum(1 for c in pw if "!@#$^*()-_=+[]{}|;:,.<>?" in c or c in "!@#$^*()-_=+[]{}|;:,.<>?") >= 1)
        check(f"Password #{i+1} no aaa",
              not any(pw[j] == pw[j+1] == pw[j+2] for j in range(len(pw)-2)))
        check(f"Password #{i+1} no user",
              "testadmin" not in pw.lower() and "nimdatset" not in pw.lower())

    # Test 5: CLI INT
    params_int = process_data({
        "ritm": "RITM00001359407",
        "minsal": "0806819-100",
        "rut": "12994496-k",
        "tipo": "INT",
        "vencimiento": "2027-02-19",
    })
    cli_int = build_cli_int(params_int)
    check("CLI INT contiene configure",
          "configure" in cli_int)
    check("CLI INT contiene user-group",
          "user-group" in cli_int)
    check("CLI INT NO contiene rulebase",
          "rulebase" not in cli_int)

    # Test 6: CLI EXT
    params_ext = process_data({
        "ritm": "RITM00001359407",
        "minsal": "0806819-100",
        "rut": "12994496-k",
        "tipo": "EXT",
        "vencimiento": "2027-02-19",
        "ips": ["10.8.110.211", "10.8.110.212"],
        "puertos": ["3389/TCP", "443/TCP"],
    })
    cli_ext = build_cli_ext(params_ext)
    check("CLI EXT contiene rulebase security rules",
          "rulebase security rules" in cli_ext)
    check("CLI EXT contiene schedule",
          "schedule" in cli_ext)
    check("CLI EXT contiene address object",
          "10.8.110.211_32" in cli_ext)
    check("CLI EXT contiene service object",
          "3389_TCP" in cli_ext)

    # Test 7: Parseo de ticket
    ticket_sample = """
    Solicitud RITM00001359407 para VPN EXT.
    Código MINSAL: 0806819-100
    RUT: 12.994.496-K
    Vencimiento: 2027-02-19
    IPs: 10.8.110.211, 10.8.110.212
    Puertos: 3389/TCP, 443/TCP
    """
    parsed = parse_ticket(ticket_sample)
    check("Ticket: RITM encontrado",
          parsed.get("ritm") == "RITM00001359407",
          f"obtenido: {parsed.get('ritm')}")
    check("Ticket: MINSAL encontrado",
          "0806819-100" in parsed.get("minsal", ""),
          f"obtenido: {parsed.get('minsal')}")
    check("Ticket: RUT encontrado",
          parsed.get("rut") is not None,
          f"obtenido: {parsed.get('rut')}")
    check("Ticket: tipo EXT encontrado",
          parsed.get("tipo") == "EXT",
          f"obtenido: {parsed.get('tipo')}")

    # Resumen
    print(f"\n{'═' * 50}")
    print(f"  Resultados: {passed}/{total} pasaron, {failed} fallaron")
    print(f"{'═' * 50}\n")

    return failed == 0


# ═══════════════════════════════════════════════════════════════════════
# MAIN / ARGPARSE
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Automatización VPN SSL GlobalProtect (INT/EXT)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:

  # Modo 2 — Parámetros INT:
  python vpn_globalprotect.py params --ritm RITM00001359407 --minsal 0806819-100 --rut 12994496-k --tipo INT --vencimiento 2027-02-19

  # Modo 2 — Parámetros EXT:
  python vpn_globalprotect.py params --ritm RITM00001359407 --minsal 0806819-100 --rut 12994496-k --tipo EXT --vencimiento 2027-02-19 --ips "10.8.110.211,10.8.110.212" --puertos "3389/TCP,443/TCP"

  # Modo 1 — Ticket desordenado:
  python vpn_globalprotect.py ticket "Solicitud RITM00001359407 VPN EXT, MINSAL 0806819-100, RUT 12.994.496-K, vence 2027-02-19, IP 10.8.110.211 puerto 3389/TCP"

  # Modo 3 — Desde Excel:
  python vpn_globalprotect.py excel PlantillaVPN.xlsx --row 3

  # Generar plantilla Excel:
  python vpn_globalprotect.py template PlantillaVPN.xlsx

  # Auto-tests:
  python vpn_globalprotect.py --test
        """
    )

    parser.add_argument("--test", action="store_true", help="Ejecutar auto-tests")
    parser.add_argument("--export-csv", type=str, help="Exportar resultado a CSV")
    parser.add_argument("--export-excel", type=str, help="Exportar resultado a Excel")

    subparsers = parser.add_subparsers(dest="mode", help="Modo de operación")

    # ── Subcomando: ticket ──
    p_ticket = subparsers.add_parser("ticket", help="Modo 1: Parsear ticket desordenado")
    p_ticket.add_argument("texto", type=str, help="Texto del ticket")

    # ── Subcomando: params ──
    p_params = subparsers.add_parser("params", help="Modo 2: Parámetros mínimos")
    p_params.add_argument("--ritm", required=True, help="RITM (ej: RITM00001359407)")
    p_params.add_argument("--minsal", required=True, help="Código MINSAL (ej: 0806819-100)")
    p_params.add_argument("--rut", required=True, help="RUT (ej: 12994496-k)")
    p_params.add_argument("--tipo", required=True, choices=["INT", "EXT", "int", "ext"],
                          help="Tipo VPN: INT o EXT")
    p_params.add_argument("--vencimiento", required=True, help="Fecha vencimiento (YYYY-MM-DD)")
    p_params.add_argument("--ips", type=str, default="", help="IPs destino separadas por coma (EXT)")
    p_params.add_argument("--puertos", type=str, default="", help="Puertos separados por coma (EXT)")
    p_params.add_argument("--obs", type=str, default="", help="Observación")
    p_params.add_argument("--vsys", default=DEFAULTS["vsys"])
    p_params.add_argument("--zona-origen", default=DEFAULTS["zona_origen"])
    p_params.add_argument("--zona-destino", default=DEFAULTS["zona_destino"])
    p_params.add_argument("--grupo-int", default=DEFAULTS["grupo_int"])
    p_params.add_argument("--grupo-ext", default=DEFAULTS["grupo_ext"])
    p_params.add_argument("--log-setting", default=DEFAULTS["log_setting"])
    p_params.add_argument("--profile-group", default=DEFAULTS["profile_group"])
    p_params.add_argument("--crear-objetos", action="store_true", default=True)
    p_params.add_argument("--no-crear-objetos", dest="crear_objetos", action="store_false")
    p_params.add_argument("--commit", action="store_true", default=False, help="Incluir commit")

    # ── Subcomando: excel ──
    p_excel = subparsers.add_parser("excel", help="Modo 3: Leer desde Excel")
    p_excel.add_argument("archivo", type=str, help="Ruta al archivo Excel")
    p_excel.add_argument("--row", type=int, default=3, help="Fila a leer (default: 3, primera de datos)")

    # ── Subcomando: template ──
    p_template = subparsers.add_parser("template", help="Generar plantilla Excel vacía")
    p_template.add_argument("archivo", type=str, help="Ruta de salida para la plantilla")

    args = parser.parse_args()

    # ── Tests ──
    if args.test:
        success = run_tests()
        return

    # ── Procesar datos según modo ──
    data = {}

    if args.mode == "ticket":
        data = parse_ticket(args.texto)
        print("\n📝 Datos extraídos del ticket:")
        for k, v in data.items():
            print(f"   {k}: {v}")
        # Verificar datos faltantes
        missing = [f for f in ["ritm", "minsal", "rut", "tipo"] if not data.get(f)]
        if missing:
            print(f"\n⚠️  Datos faltantes: {', '.join(missing)}")
            print("   Proporciona los datos faltantes usando modo params.")
            return

    elif args.mode == "params":
        data = {
            "ritm": args.ritm,
            "minsal": args.minsal,
            "rut": args.rut,
            "tipo": args.tipo.upper(),
            "vencimiento": args.vencimiento,
            "ips": [ip.strip() for ip in args.ips.split(",") if ip.strip()] if args.ips else [],
            "puertos": [p.strip() for p in args.puertos.split(",") if p.strip()] if args.puertos else [],
            "obs": args.obs,
            "vsys": args.vsys,
            "zona_origen": args.zona_origen,
            "zona_destino": args.zona_destino,
            "grupo_int": args.grupo_int,
            "grupo_ext": args.grupo_ext,
            "log_setting": args.log_setting,
            "profile_group": args.profile_group,
            "crear_objetos": args.crear_objetos,
            "incluir_commit": args.commit,
        }

    elif args.mode == "excel":
        data = read_excel_row(args.archivo, args.row)
        print(f"\n📊 Datos leídos de Excel (fila {args.row}):")
        for k, v in data.items():
            print(f"   {k}: {v}")

    else:
        parser.print_help()
        return

    # ── Procesar y renderizar ──
    try:
        params = process_data(data)
        output = render_output(params)
        print(output)

        # ── Exports opcionales ──
        if args.export_csv:
            export_csv(params, args.export_csv)
        if args.export_excel:
            export_excel_output(params, args.export_excel)

    except ValueError as e:
        print(f"\n❌ Error de validación:\n   {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
